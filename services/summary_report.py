"""
Summary_report.py
-----------------
Generates a polished HTML summary report strictly from the finalised Excel
report file (source of truth).  Never reads Validation_Output directly.

Public API
----------
    summary_report_writer(excel_path: str) -> str
        Reads the Excel file at *excel_path*, builds the HTML summary, writes
        it to the same directory, and returns the HTML file path.

Filename format
---------------
    Summary-report_<json_name>_<excel_report_basename>.html
"""

import ast
import base64
import io
import os
import re
import html as _html
from collections import defaultdict, OrderedDict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment

#from Input import JSON_URL

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_STATUS_NORM = {
    'passed': 'Passed', 'passes': 'Passed',
    'failed': 'Failed',
    'not tested': 'Not Tested',
    'observation': 'Observation',
    'not applicable': 'Not Applicable',
    'n/a': 'Not Applicable',
    'na': 'Not Applicable',
}


def _norm(status):
    return _STATUS_NORM.get(str(status).strip().lower(), str(status).strip())


def _module_overall(rows):
    """
    Priority:
        Failed            – any Failed row present
        Passed with caveats – no Failed, but Not Tested or Observation present
        Passed            – all rows are Passed
    """
    statuses = {_norm(r['Status']) for r in rows}
    if 'Failed' in statuses:
        return 'Failed'
    if 'Not Tested' in statuses or 'Observation' in statuses or 'Not Applicable' in statuses:
        return 'Passed with caveats'
    return 'Passed'


def _cls(status):
    return {
        'Passed':             'passed',
        'Failed':             'failed',
        'Not Tested':         'nt',
        'Observation':        'obs',
        'Not Applicable':     'na',
        'Passed with caveats': 'caveats',
    }.get(status, 'nt')


def _esc(v):
    return _html.escape(str(v)) if v is not None else ''


def _truncate(text, limit=400):
    t = str(text).strip()
    return t[:limit] + ' …' if len(t) > limit else t


#def _json_name():
    """Derive a short identifier from the JSON_URL (or file path)."""
    #try:
        #return os.path.splitext(os.path.basename(JSON_URL.rstrip('/')))[0]
    #except Exception:
        #return 'feed'


# ---------------------------------------------------------------------------
# CSS  (plain string — kept separate to avoid f-string brace-escaping issues)
# ---------------------------------------------------------------------------

_CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto,
               'Helvetica Neue', Arial, sans-serif;
  background: #F1F5F9; color: #1E293B; line-height: 1.6; font-size: 14px;
}

/* ── Hero ── */
.hero {
  background: linear-gradient(140deg, #0F172A 0%, #1E293B 55%, #1a3a5c 100%);
  padding: 52px 48px 44px; color: #fff;
}
.hero-inner        { max-width: 1080px; margin: 0 auto; }
.hero-eyebrow      { font-size: 11px; text-transform: uppercase; letter-spacing: 2px;
                     color: #64748B; margin-bottom: 12px; font-weight: 600; }
.hero-title        { font-size: 32px; font-weight: 800; letter-spacing: -.5px; }
.hero-meta         { font-size: 13px; color: #94A3B8; margin-top: 10px;
                     display: flex; flex-wrap: wrap; gap: 20px; align-items: center; }
.hero-meta span    { display: flex; align-items: center; gap: 6px; }
.hero-pill         { display: inline-flex; align-items: center; gap: 9px;
                     margin-top: 22px; border-radius: 30px; padding: 8px 20px;
                     font-size: 13px; font-weight: 700; letter-spacing: .2px;
                     border: 1px solid transparent; }
.hero-dot          { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }

/* ── Layout ── */
.container { max-width: 1080px; margin: 0 auto; padding: 36px 24px 52px; }

/* ── Base card ── */
.card {
  background: #fff; border-radius: 14px; padding: 26px 30px;
  box-shadow: 0 1px 2px rgba(0,0,0,.06), 0 6px 24px rgba(0,0,0,.06);
  margin-bottom: 24px;
}
.card-label { font-size: 11px; font-weight: 700; text-transform: uppercase;
              letter-spacing: 1.2px; color: #94A3B8; margin-bottom: 18px; }

/* ── Inputs ── */
.input-grid              { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 24px; }
.input-item label        { display: block; font-size: 11px; text-transform: uppercase;
                           letter-spacing: .8px; color: #94A3B8; font-weight: 700; }
.input-item .val         { display: block; font-size: 13px; font-weight: 500; color: #1E293B;
                           margin-top: 5px; word-break: break-all; }

/* ── KPI row ── */
.kpi-grid    { display: grid; grid-template-columns: repeat(5, 1fr); gap: 14px;
               margin-bottom: 32px; }
.kpi-card    { background: #fff; border-radius: 14px; padding: 22px 14px 20px; text-align: center;
               box-shadow: 0 1px 2px rgba(0,0,0,.06), 0 5px 18px rgba(0,0,0,.06); }
.kpi-value   { font-size: 40px; font-weight: 800; line-height: 1.1; }
.kpi-label   { font-size: 10px; text-transform: uppercase; letter-spacing: 1px;
               color: #64748B; margin-top: 7px; font-weight: 700; }
.kpi-sub     { font-size: 11px; color: #94A3B8; margin-top: 3px; }

/* ── Status badges ── */
.badge         { display: inline-block; padding: 3px 11px; border-radius: 20px;
                 font-size: 11px; font-weight: 700; letter-spacing: .3px;
                 border: 1px solid transparent; white-space: nowrap; }
.badge-passed  { color: #166534; background: #DCFCE7; border-color: #86EFAC; }
.badge-failed  { color: #991B1B; background: #FEE2E2; border-color: #FCA5A5; }
.badge-nt      { color: #1E40AF; background: #DBEAFE; border-color: #93C5FD; }
.badge-obs     { color: #854D0E; background: #FEF9C3; border-color: #FDE68A; }
.badge-na      { color: #475569; background: #F1F5F9; border-color: #CBD5E1; }
.badge-caveats { color: #92400E; background: #FEF3C7; border-color: #FCD34D; }

/* ── Module cards ── */
.module-card          { background: #fff; border-radius: 14px; overflow: hidden;
                        margin-bottom: 14px;
                        box-shadow: 0 1px 2px rgba(0,0,0,.06), 0 5px 18px rgba(0,0,0,.06); }
.module-card.bar-passed  { border-left: 5px solid #22C55E; }
.module-card.bar-failed  { border-left: 5px solid #EF4444; }
.module-card.bar-nt      { border-left: 5px solid #3B82F6; }
.module-card.bar-obs     { border-left: 5px solid #F59E0B; }
.module-card.bar-caveats { border-left: 5px solid #F59E0B; }

.module-header  { display: flex; align-items: center; justify-content: space-between;
                  padding: 18px 24px; border-bottom: 1px solid #F8FAFC; }
.module-name    { font-size: 15px; font-weight: 700; }
.module-meta-r  { display: flex; align-items: center; gap: 12px; }
.check-count    { font-size: 12px; color: #94A3B8; font-weight: 500; }

.module-body    { padding: 18px 24px 20px; }

/* ── Stat pills ── */
.stats-row  { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 14px; }
.stat-pill  { display: inline-flex; align-items: center; gap: 6px;
              padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 700; }
.stat-dot   { width: 7px; height: 7px; border-radius: 50%; flex-shrink: 0; }

/* ── Progress bar ── */
.prog-wrap  { background: #F1F5F9; border-radius: 8px; height: 5px; overflow: hidden; }
.prog-fill  { height: 100%; border-radius: 8px; }
.prog-label { font-size: 11px; color: #94A3B8; text-align: right; margin-top: 5px; }

/* ── Scenario details ── */
details         { margin-top: 12px; border-radius: 8px; overflow: hidden; }
details summary { cursor: pointer; font-size: 11px; font-weight: 700;
                  text-transform: uppercase; letter-spacing: .7px; padding: 9px 14px;
                  display: flex; align-items: center; gap: 8px;
                  user-select: none; list-style: none; }
details summary::-webkit-details-marker { display: none; }
.sum-failed { background: #FEF2F2; color: #991B1B; }
.sum-obs    { background: #FFFBEB; color: #92400E; }
.sum-nt     { background: #EFF6FF; color: #1E40AF; }
.chevron    { font-size: 9px; transition: transform .18s ease; flex-shrink: 0; }
details[open] .chevron { transform: rotate(90deg); }

.sc-list         { list-style: none; padding: 0 0 4px 0; }
.sc-list li      { padding: 9px 16px 9px 14px; font-size: 13px; border-left: 3px solid; }
.sc-list li + li { border-top: 1px solid transparent; }
.sc-list li.fail { background: #FFF8F8; border-color: #FCA5A5; }
.sc-list li.obs  { background: #FFFEF5; border-color: #FDE68A; }
.sc-list li.nt   { background: #F5F9FF; border-color: #93C5FD; }
.sc-name   { font-weight: 600; color: #1E293B; }
.sc-result { font-size: 11.5px; color: #64748B; margin-top: 3px;
             word-break: break-word; white-space: pre-wrap; line-height: 1.5; }

/* ── Failure Summary ── */
.fs-list         { margin-top: 4px; }
.fs-item         { border-radius: 8px; overflow: hidden; margin-bottom: 6px; }
.fs-summary      { display: flex; align-items: center; gap: 8px; padding: 10px 14px;
                   background: #FFF8F8; cursor: pointer; user-select: none; list-style: none; }
.fs-summary::-webkit-details-marker { display: none; }
.fs-sc-name      { flex: 1; min-width: 0; font-size: 13px; font-weight: 600; color: #1E293B; }
.fs-id-badge     { flex-shrink: 0; background: #FEE2E2; color: #991B1B;
                   border: 1px solid #FCA5A5; border-radius: 20px;
                   padding: 2px 10px; font-size: 11px; font-weight: 700; white-space: nowrap; }
.fs-ids-body     { padding: 8px 16px 10px 36px; font-size: 12px; color: #475569;
                   background: #FFF2F2; border-left: 3px solid #FCA5A5;
                   word-break: break-all; line-height: 1.6; }
.fs-none         { font-size: 13px; color: #94A3B8; font-style: italic; }

/* ── Failure Summary table ── */
.fs-table          { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 4px;
                     table-layout: fixed; }
.fs-table th       { padding: 9px 12px; text-align: left; font-size: 11px; font-weight: 700;
                     text-transform: uppercase; letter-spacing: .8px; color: #64748B;
                     background: #F8FAFC; border-bottom: 2px solid #E2E8F0; }
.fs-table th.fs-col-id     { text-align: center; width: 18%; }
.fs-table th.fs-col-field  { width: 14%; }
.fs-table th.fs-col-issues { width: 68%; }
.fs-table td       { padding: 9px 12px; border-bottom: 1px solid #F1F5F9; vertical-align: top;
                     overflow-wrap: anywhere; }
.fs-table td.fs-col-id    { text-align: left; font-weight: 700; color: #991B1B;
                             font-size: 12px; font-family: 'Courier New', monospace;
                             word-break: break-all; white-space: normal; max-width: 0; }
.fs-table td.fs-col-field { color: #475569; font-weight: 600;
                             word-break: break-word; max-width: 0; }
.fs-table td.fs-col-issues { color: #1E293B; word-break: break-word; white-space: normal;
                             line-height: 1.6; max-width: 0; }
.fs-table tr:last-child td { border-bottom: none; }
.fs-table tbody tr:hover td { background: #FFF8F8; }

/* ── Coverage card ── */
.cov-grid   { display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }
.cov-section { }
.cov-title  { font-size: 12px; font-weight: 700; text-transform: uppercase;
               letter-spacing: .8px; color: #475569; margin-bottom: 12px;
               display: flex; align-items: center; gap: 8px; }
.cov-title-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
.cov-table  { width: 100%; border-collapse: collapse; font-size: 12px; }
.cov-table td { padding: 5px 8px; border-bottom: 1px solid #F1F5F9; }
.cov-table td:first-child { color: #64748B; font-weight: 500; width: 60%; }
.cov-table td:last-child  { font-weight: 700; text-align: right; color: #1E293B; }
.cov-reasons { margin-top: 10px; }
.cov-reasons-label { font-size: 11px; font-weight: 600; color: #94A3B8;
                     text-transform: uppercase; letter-spacing: .6px; margin-bottom: 6px; }
.cov-reason-pill { display: inline-flex; align-items: center; gap: 5px; margin: 2px 3px 2px 0;
                   padding: 3px 10px; border-radius: 20px; font-size: 11px; font-weight: 600;
                   background: #FEE2E2; color: #991B1B; border: 1px solid #FCA5A5; }
.cov-skip-pill  { background: #FEF9C3; color: #854D0E; border-color: #FDE68A; }
.cov-none { font-size: 12px; color: #94A3B8; font-style: italic; }

/* ── Open Report button ── */
.open-report-btn {
  display: inline-flex; align-items: center; gap: 8px;
  padding: 10px 22px; background: #1E293B; color: #fff !important;
  border-radius: 10px; font-size: 13px; font-weight: 700; letter-spacing: .3px;
  text-decoration: none; box-shadow: 0 2px 8px rgba(0,0,0,.18);
  transition: background .15s ease, box-shadow .15s ease;
}
.open-report-btn:hover { background: #334155; box-shadow: 0 4px 14px rgba(0,0,0,.22); }
.btn-row { text-align: right; margin-bottom: 22px; }

/* ── Module blurb ── */
.mod-blurb {
  background: #F8FAFC; border-radius: 8px; padding: 11px 14px;
  margin-bottom: 14px; font-size: 13px; line-height: 1.65;
  border: 1px solid #E2E8F0;
}
.mod-blurb-item { display: block; }
.mod-blurb-item + .mod-blurb-item { margin-top: 4px; }

/* ── Section heading ── */
.section-head { font-size: 11px; font-weight: 700; text-transform: uppercase;
                letter-spacing: 1.2px; color: #94A3B8;
                margin-bottom: 14px; margin-top: 4px; }

/* ── Footer ── */
.footer { text-align: center; padding: 36px 24px 28px; color: #94A3B8; font-size: 12px;
          border-top: 1px solid #E2E8F0; margin-top: 8px; }
.footer strong { color: #64748B; }

@media (max-width: 720px) {
  .kpi-grid   { grid-template-columns: repeat(2, 1fr); }
  .input-grid { grid-template-columns: 1fr; }
  .hero       { padding: 32px 20px 28px; }
  .container  { padding: 20px 16px; }
}

/* ── Download button ── */
.dl-btn { display:inline-block; margin-left:14px; padding:4px 14px; font-size:12px;
          font-weight:600; color:#fff; background:#3B82F6; border:none; border-radius:6px;
          cursor:pointer; vertical-align:middle; line-height:1.6; }
.dl-btn:hover { background:#2563EB; }
.dl-btn-tab {
  margin-left: 0; padding: 7px 14px; font-size: 12px; white-space: nowrap;
}
.tab-bar-dl-hidden { display: none !important; }

/* ── Tabs ── */
.tab-bar {
  display: flex; flex-wrap: wrap; align-items: center; justify-content: space-between;
  gap: 8px 12px;
  background: #fff; border-radius: 14px 14px 0 0; padding: 8px 12px 0;
  box-shadow: 0 1px 2px rgba(0,0,0,.06), 0 6px 24px rgba(0,0,0,.06);
  border-bottom: 1px solid #E2E8F0; margin-bottom: 0;
}
.tab-bar-tabs {
  display: flex; flex-wrap: wrap; gap: 4px; min-width: 0; flex: 1 1 auto;
}
.tab-bar-actions {
  display: flex; align-items: center; gap: 8px;
  flex: 0 0 auto; padding-bottom: 8px;
}
.tab-btn {
  appearance: none; border: none; background: transparent;
  padding: 12px 18px; font-size: 13px; font-weight: 700; color: #64748B;
  cursor: pointer; border-radius: 10px 10px 0 0; letter-spacing: .2px;
  border-bottom: 3px solid transparent; margin-bottom: -1px;
}
.tab-btn:hover { color: #1E293B; background: #F8FAFC; }
.tab-btn.active {
  color: #0F172A; background: #F8FAFC;
  border-bottom-color: #3B82F6;
}
.tab-panels {
  background: #fff; border-radius: 0 0 14px 14px; padding: 22px 24px 26px;
  box-shadow: 0 1px 2px rgba(0,0,0,.06), 0 6px 24px rgba(0,0,0,.06);
  margin-bottom: 24px;
}
.tab-panel { display: none; }
.tab-panel.active { display: block; }
.tab-panels .card {
  box-shadow: none; margin-bottom: 0; padding: 0;
  border-radius: 0;
}

/* ── Complete test cases table ── */
.tc-kpi-row {
  display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px;
}
.tc-table-wrap {
  overflow: auto; border: 1px solid #E2E8F0; border-radius: 10px;
  max-height: 640px;
}
.tc-table {
  width: max-content; border-collapse: separate; border-spacing: 0;
  font-size: 13px; table-layout: fixed;
}
.tc-table th {
  position: sticky; top: 0; z-index: 2;
  padding: 8px 10px 8px 6px; text-align: left; font-size: 11px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .7px; color: #64748B;
  background: #F8FAFC; border-bottom: 2px solid #E2E8F0;
  box-sizing: border-box; white-space: nowrap;
  /* max-width:0 enables shrink; overflow must stay visible so resize grip is clickable */
  max-width: 0;
}
.tc-th-inner {
  position: relative; display: block; width: 100%; min-height: 1.2em;
  overflow: visible;
}
.tc-th-label {
  display: block; overflow: hidden; text-overflow: ellipsis; padding-right: 10px;
}
.tc-col-resizer {
  position: absolute; top: -16px; right: -4px; bottom: -16px; width: 18px;
  cursor: col-resize; user-select: none; z-index: 20; touch-action: none;
}
.tc-table th:last-child .tc-col-resizer {
  right: -4px; width: 20px;
}
.tc-col-resizer::after {
  content: ''; position: absolute; top: 18%; bottom: 18%; left: 50%;
  width: 2px; margin-left: -1px; background: #94A3B8; border-radius: 1px;
}
.tc-col-resizer:hover::after, .tc-col-resizer.resizing::after {
  background: #3B82F6; width: 3px; margin-left: -1.5px;
}
body.tc-col-resizing, body.tc-col-resizing * {
  cursor: col-resize !important; user-select: none !important;
}
body.tc-col-resizing { -webkit-user-select: none; }
.tc-table td {
  padding: 8px 6px; border-bottom: 1px solid #F1F5F9; vertical-align: top;
  overflow-wrap: anywhere; word-break: break-word; color: #1E293B;
  box-sizing: border-box; overflow: hidden; max-width: 0;
}
.tc-table tr:last-child td { border-bottom: none; }
.tc-table tbody tr:hover td { background: #F8FAFC; }
.tc-empty { font-size: 13px; color: #94A3B8; font-style: italic; padding: 12px 0; }
.tab-blurb {
  background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 8px;
  padding: 11px 14px; margin-bottom: 14px; font-size: 13px; line-height: 1.65; color: #475569;
}
.tc-filter-pill {
  appearance: none; border: 1px solid transparent; cursor: pointer;
  display: inline-flex; align-items: center; gap: 6px;
  padding: 5px 12px; border-radius: 20px; font-size: 12px; font-weight: 700;
}
.tc-filter-pill.active-filter {
  outline: 2px solid #1E293B; outline-offset: 1px;
}
.tc-filter-pill:hover { filter: brightness(0.97); }
.tc-assets-cell { line-height: 1.5; min-width: 0; }
.tc-assets-preview { word-break: break-word; overflow: hidden; }
.tc-assets-full { display: none; word-break: break-word; white-space: pre-wrap; margin-top: 4px; }
.tc-assets-cell.expanded .tc-assets-full { display: block; }
.tc-assets-cell.expanded .tc-assets-preview { display: none; }
.tc-assets-actions { margin-top: 6px; display: flex; flex-wrap: wrap; gap: 8px; }
.tc-assets-actions.tc-actions-hidden { display: none !important; }
.tc-more-btn, .tc-copy-btn {
  appearance: none; border: none; background: none; padding: 0;
  font-size: 11px; font-weight: 700; color: #2563EB; cursor: pointer;
  text-decoration: underline;
}
.tc-more-btn:hover, .tc-copy-btn:hover { color: #1D4ED8; }
.tc-table tbody tr.tc-row-hidden { display: none; }

/* ── Grouped failed cases (by Issue Summary) ── */
.gf-item { border: 1px solid #E2E8F0; border-radius: 10px; margin-bottom: 8px; overflow: hidden; }
.gf-summary {
  display: flex; align-items: center; gap: 10px; padding: 12px 14px;
  background: #FFF8F8; cursor: pointer; user-select: none; list-style: none;
}
.gf-summary::-webkit-details-marker { display: none; }
.gf-sc-name { flex: 1; min-width: 0; font-size: 13px; font-weight: 600; color: #1E293B; }
.gf-count {
  flex-shrink: 0; background: #FEE2E2; color: #991B1B; border: 1px solid #FCA5A5;
  border-radius: 20px; padding: 2px 10px; font-size: 11px; font-weight: 700;
}
.gf-body { padding: 0; background: #fff; border-top: 1px solid #FEE2E2; }
.gf-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.gf-table th {
  padding: 8px 12px; text-align: left; font-size: 10px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .7px; color: #64748B;
  background: #F8FAFC; border-bottom: 1px solid #E2E8F0;
}
.gf-table td {
  padding: 8px 12px; border-bottom: 1px solid #F1F5F9; vertical-align: top;
  overflow-wrap: anywhere; color: #334155;
}
.gf-table tr:last-child td { border-bottom: none; }
.gf-asset-id {
  font-family: 'Courier New', monospace; font-size: 12px; font-weight: 700;
  color: #991B1B; word-break: break-all;
}
.gf-empty { font-size: 13px; color: #94A3B8; font-style: italic; }
.gf-assets-wrap { padding: 0; }
.gf-assets-header {
  padding: 8px 12px; font-size: 10px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .7px; color: #64748B;
  background: #F8FAFC; border-bottom: 1px solid #E2E8F0;
}
.gf-assets-wrap .tc-assets-cell { padding: 10px 12px; }
.gf-assets-wrap .gf-asset-id { white-space: pre-wrap; }
"""

# ---------------------------------------------------------------------------
# HTML component builders
# ---------------------------------------------------------------------------

# def _kpi_card(value, label, color, border_top, sub=''):
#     sub_html = f'<div class="kpi-sub">{_esc(sub)}</div>' if sub else ''
#     return (
#         f'<div class="kpi-card" style="border-top:3px solid {border_top};">'
#         f'<div class="kpi-value" style="color:{color};">{value}</div>'
#         f'<div class="kpi-label">{label}</div>'
#         f'{sub_html}'
#         f'</div>'
#     )


def _stat_pill(count, label, dot, text, bg):
    return (
        f'<span class="stat-pill" style="background:{bg};color:{text};">'
        f'<span class="stat-dot" style="background:{dot};"></span>'
        f'{count} {label}</span>'
    )


def _scenario_li(row, css_class):
    actual = _truncate(row.get('Actual Result', '') or '')
    actual_html = f'<div class="sc-result">{_esc(actual)}</div>' if actual else ''
    return (
        f'<li class="{css_class}">'
        f'<div class="sc-name">{_esc(row.get("Scenario", ""))}</div>'
        f'{actual_html}'
        f'</li>'
    )


def _detail_section(items, label, css_class, sum_cls, open_attr=''):
    if not items:
        return ''
    lis = ''.join(_scenario_li(r, css_class) for r in items)
    return (
        f'<details {open_attr}>'
        f'<summary class="{sum_cls}">'
        f'<span class="chevron">&#9658;</span>'
        f'&nbsp;{len(items)} {label}'
        f'</summary>'
        f'<ul class="sc-list">{lis}</ul>'
        f'</details>'
    )


def _module_blurb(rows, overall, mc):
    """Short stakeholder-friendly summary shown at the top of each module body."""
    parts = []

    if mc['Failed'] > 0:
        parts.append(
            f'<span class="mod-blurb-item" style="color:#991B1B;font-weight:700;">'
            f'&#10007;&nbsp;{mc["Failed"]} check(s) failed</span>'
        )

    if overall == 'Passed with caveats':
        caveat_parts = []
        if mc['Not Tested'] > 0:
            caveat_parts.append(f'{mc["Not Tested"]} not tested')
        if mc['Observation'] > 0:
            caveat_parts.append(f'{mc["Observation"]} observation(s)')
        parts.append(
            f'<span class="mod-blurb-item" style="color:#92400E;font-weight:600;">'
            f'&#9432;&nbsp;{", ".join(caveat_parts)} &mdash; optional checks, no blocking failures.</span>'
        )

    if overall == 'Passed':
        parts.append(
            '<span class="mod-blurb-item" style="color:#166534;font-weight:600;">'
            '&#10003;&nbsp;All checks passed.</span>'
        )

    # Grouped not-tested reasons (top 3 unique patterns)
    nt_rows = [r for r in rows if _norm(r['Status']) == 'Not Tested']
    if nt_rows:
        reason_counts = defaultdict(int)
        for r in nt_rows:
            raw = str(r.get('Actual Result', '') or '').strip()
            # Use first 80 chars as the reason key to group similar messages
            key = raw[:80] + ('…' if len(raw) > 80 else '')
            if key:
                reason_counts[key] += 1
        if reason_counts:
            top = sorted(reason_counts.items(), key=lambda x: -x[1])[:3]
            reasons_html = '&nbsp;&middot;&nbsp;'.join(
                f'{_esc(r)} <em>({c})</em>' for r, c in top
            )
            parts.append(
                f'<span class="mod-blurb-item" style="font-size:12px;color:#64748B;">'
                f'&#9702;&nbsp;Not-tested reasons: {reasons_html}</span>'
            )

    if not parts:
        return ''
    return '<div class="mod-blurb">' + ''.join(parts) + '</div>'


def _module_card(mod_name, rows):
    overall  = _module_overall(rows)
    bar_cls  = f'bar-{_cls(overall)}'

    mc = defaultdict(int)
    for r in rows:
        mc[_norm(r['Status'])] += 1

    pills = (
        _stat_pill(mc['Passed'],     'Passed',      '#22C55E', '#166534', '#DCFCE7') +
        _stat_pill(mc['Failed'],     'Failed',      '#EF4444', '#991B1B', '#FEE2E2') +
        _stat_pill(mc['Not Tested'], 'Not Tested',  '#3B82F6', '#1E40AF', '#DBEAFE') +
        _stat_pill(mc['Observation'], 'Observation','#F59E0B', '#854D0E', '#FEF9C3')
    )

    fail_rows = [r for r in rows if _norm(r['Status']) == 'Failed']
    obs_rows  = [r for r in rows if _norm(r['Status']) == 'Observation']
    nt_rows   = [r for r in rows if _norm(r['Status']) == 'Not Tested']

    details_html = (
        _detail_section(fail_rows, 'Failed',       'fail', 'sum-failed', 'open') +
        _detail_section(obs_rows,  'Observations', 'obs',  'sum-obs') +
        _detail_section(nt_rows,   'Not Tested',   'nt',   'sum-nt')
    )

    pass_pct   = int(mc['Passed'] / len(rows) * 100) if rows else 0
    prog_color = '#22C55E' if pass_pct >= 80 else ('#F59E0B' if pass_pct >= 50 else '#EF4444')

    blurb = _module_blurb(rows, overall, mc)

    return (
        f'<div class="module-card {bar_cls}">'
        f'<div class="module-header">'
        f'<span class="module-name">{_esc(mod_name)}</span>'
        f'<div class="module-meta-r">'
        f'<span class="check-count">{len(rows)} checks</span>'
        f'<span class="badge badge-{_cls(overall)}">{_esc(overall)}</span>'
        f'</div>'
        f'</div>'
        f'<div class="module-body">'
        f'{blurb}'
        f'<div class="stats-row">{pills}</div>'
        f'<div class="prog-wrap">'
        f'<div class="prog-fill" style="width:{pass_pct}%;background:{prog_color};"></div>'
        f'</div>'
        f'<div class="prog-label">{pass_pct}% passed</div>'
        f'{details_html}'
        f'</div>'
        f'</div>'
    )

# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def _read_excel(excel_path):
    """Read the Excel report and return a list of row dicts."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        d = {headers[i]: (row[i] if row[i] is not None else '') for i in range(len(headers))}
        data.append(d)
    return data

# ---------------------------------------------------------------------------
# Failure Summary helpers
# ---------------------------------------------------------------------------

_ID_SKIP_WORDS = frozenset([
    'true', 'false', 'none', 'null', 'nan', 'ok', 'not', 'tested',
    'passed', 'failed', 'observation', 'error', 'type', 'format',
    'string', 'status', 'code', 'available', '',
])


def _extract_ids(actual_result):
    """
    Conservative extraction of asset IDs from an Actual Result string.

    Strategies (in order):
      1. Content inside {} and [] — handles set literals and dict key-value pairs.
         For dict-like items (containing ':'), only the part before ':' is kept.
      2. Explicit  id=<value>  or  asset_id=<value>  patterns anywhere in the text.

    Returns a set of ID strings (deduplicated).
    """
    if not actual_result:
        return set()

    text = str(actual_result)
    ids = set()

    # Strategy 1: extract items from {} / [] blocks
    for block_m in re.finditer(r'[{\[](.*?)[}\]]', text, re.DOTALL):
        block_content = block_m.group(1)
        for token in block_content.split(','):
            token = token.strip()
            # Dict entry — take the key (before the first ':')
            if ':' in token:
                token = token.split(':', 1)[0].strip()
            # Strip surrounding quotes and whitespace
            token = token.strip("'\" \t\n\r")
            if token and token.lower() not in _ID_SKIP_WORDS:
                if re.match(r'^[A-Za-z0-9][\w\-]*$', token):
                    ids.add(token)

    # Strategy 2: explicit id= / asset_id= patterns
    for m in re.finditer(
        r'\b(?:asset_id|id)\s*=\s*[\'"]?([A-Za-z0-9][\w\-]*)[\'"]?',
        text, re.IGNORECASE
    ):
        candidate = m.group(1)
        if candidate.lower() not in _ID_SKIP_WORDS:
            ids.add(candidate)

    return ids


# Words that appear as dict *values* in node-validation results and must never
# be mistaken for asset IDs or module labels.
_FIELD_SKIP_WORDS = frozenset([
    'rating', 'genres', 'movies', 'tvspecials', 'tvspecial',
    'series', 'episodes', 'episode', 'shortformvideos', 'shortformvideo',
    'language', 'thumbnail', 'title', 'description', 'tags',
    'type', 'format', 'status', 'code', 'available', '',
])


# ---------------------------------------------------------------------------
# Scenario exclusion list  (HTML rendering only — Excel is never modified)
# ---------------------------------------------------------------------------

_EXCLUDED_SCENARIOS = frozenset([
    # 'Validate quality of the Branded Thumbnail',
    # 'Validate Title treatment of Branded Thumbnail',
    # 'Validate Channel_Level_Title Language',
    # 'Validate Content_Level Language is matching with Channel_Level Language',
    # 'Validate Channel_Long_Description Language is matching with Channel_Level_Language',
    # 'Validate Channel_Short_Description Language is matching with Channel_Level_Language',
    # 'Validate Title treatment of Channel_Thumbnail',
    # 'Validate quality of the Episode_Thumbnail',
    # 'Validate Title treatment of Episode_Thumbnail',
    # 'Validate quality of the Series_Thumbnail',
    # 'Validate Title treatment of Series_Thumbnail',
    # 'Validate Series_Short_Description Language is matching with Channel_Level_Language',
    # 'Validate quality of the Movie_Thumbnail',
    # 'Validate Title treatment of Movie_Thumbnail',
    # 'Validate Movie_Short_Description Language is matching with Channel_Level_Language',
    # 'Validate quality of the shortFormVideos_Thumbnail',
    # 'Validate Title treatment of shortFormVideos_Thumbnail',
    # 'Validate quality of the tvSpecials_Thumbnail',
    # 'Validate Title treatment of tvSpecials_Thumbnail',
    # 'Validate tvSpecials_Short_Description Language is matching with Channel_Level_Language',
])


def _filter_rows(rows):
    """Return rows with excluded scenarios removed (HTML rendering only)."""
    return [
        r for r in rows
        if str(r.get('Scenario', '') or '').strip() not in _EXCLUDED_SCENARIOS
    ]


def _parse_asset_ids_column(text):
    """
    Parse the 'Asset IDs' column value into a set of display identifiers.

    Priority (asset IDs first, module label fallback):
      1. Real asset IDs — dict keys that contain at least one digit and match
         an alphanumeric/hyphen identifier pattern, e.g. m54745, BellMedia-2931555-S3E12.
      2. Module labels — if no real asset IDs are found, the dict key is treated
         as a module-level label (multi-word or contains "Nodes").
         Duplicate trailing "Nodes" is normalised:
           "JSON Nodes Nodes"       → "JSON Nodes"
           "Live Feeds Nodes Nodes" → "Live Feeds Nodes"

    Plain comma-separated values (no braces) are also supported as a fallback
    for simple "id1, id2" lists produced by the new reporting schema.

    Returns a set of strings (deduplicated).
    """
    if not text:
        return set()

    raw_keys = []

    # Extract top-level dict keys: the segment before the first ':' inside each
    # outermost { } that does not itself contain { or } before the colon.
    # This correctly handles both:
    #   {'m54745': ['rating', 'rating']},{'m54629': ['rating']}
    #   {JSON Nodes Nodes : {'movies', 'tvSpecials'}}
    for m in re.finditer(r'\{([^{}:]*?)\s*:', text):
        key = m.group(1).strip().strip("'\" \t{}[]")
        if key:
            raw_keys.append(key)

    # Fallback: plain comma-separated list with no brace-dict structure
    if not raw_keys and '{' not in text and '[' not in text:
        for token in re.split(r'[,\n]+', text):
            token = token.strip()
            if ':' in token:
                token = token.split(':', 1)[0].strip()
            token = token.strip("'\" \t")
            if token:
                raw_keys.append(token)

    asset_ids = []
    module_labels = []

    for key in raw_keys:
        if not key:
            continue
        key_norm = key.lower().replace(' ', '').replace('_', '')
        if key_norm in _FIELD_SKIP_WORDS or key_norm in _ID_SKIP_WORDS:
            continue

        # Real asset ID: single word (no spaces), alphanumeric/hyphens, contains a digit
        if re.match(r'^[A-Za-z0-9][\w\-]*$', key) and re.search(r'\d', key):
            asset_ids.append(key)
        # Module label: multi-word phrase or contains "Nodes" keyword
        elif ' ' in key or re.search(r'\bNodes?\b', key, re.IGNORECASE):
            label = re.sub(r'\bNodes\b\s+\bNodes\b', 'Nodes', key, flags=re.IGNORECASE).strip()
            module_labels.append(label)
        # Single-word, no digit — only include if it survived the skip-word filter
        elif re.match(r'^[A-Za-z0-9][\w\-]*$', key):
            asset_ids.append(key)

    # Priority: return real asset IDs; fall back to module labels
    result = asset_ids if asset_ids else module_labels

    # Deduplicate while preserving first-seen order, then return as a set
    seen: set = set()
    out = []
    for item in result:
        if item not in seen:
            seen.add(item)
            out.append(item)

    return set(out)


# ---------------------------------------------------------------------------
# Module label cleanup for Failure Summary column 2
# ---------------------------------------------------------------------------

_MODULE_CLEAN_MAP = [
    (re.compile(r'JSON\s+Nodes',                              re.IGNORECASE), 'JSON'),
    (re.compile(r'Live\s+Feeds?\s+Nodes',                     re.IGNORECASE), 'Live Feed'),
    (re.compile(r'Series\s+Nodes',                            re.IGNORECASE), 'Series'),
    (re.compile(r'Episodes?\s+Nodes',                         re.IGNORECASE), 'Episodes'),
    (re.compile(r'Schedule\s+Nodes',                          re.IGNORECASE), 'Schedule'),
    (re.compile(r'Content[_\s]+Nodes',                        re.IGNORECASE), 'Content'),
    (re.compile(r'Movies?\s+Nodes',                           re.IGNORECASE), 'Movies'),
    (re.compile(r'TV[_\s]+Specials?\s+Nodes',                 re.IGNORECASE), 'TV Specials'),
    (re.compile(r'Short[_\s]+Form[_\s]*Videos?\s+Nodes',      re.IGNORECASE), 'Short-Form Videos'),
    (re.compile(r'Seasons?\s+Nodes',                          re.IGNORECASE), 'Seasons'),
]


def _clean_module_label(module_raw):
    """Map a raw Module column value to a clean label for column 2 of the Failure table."""
    raw = str(module_raw or '').strip()
    for pattern, label in _MODULE_CLEAN_MAP:
        if pattern.search(raw):
            return label
    # Fallback: strip all repeated trailing "Nodes" suffixes
    return re.sub(r'(\s+Nodes)+\s*$', '', raw, flags=re.IGNORECASE).strip() or raw


def _extract_asset_ids_for_grouping(asset_ids_text):
    """
    Extract real asset IDs (must contain at least one digit) from the Asset IDs
    column value.  Returns a deduplicated ordered list, or [] when no real IDs
    are present (i.e. this is a module-level / root-level failure).

    Three-pass strategy
    -------------------
    Pass 1 — quoted dict keys (handles spaces):
        Captures the full string between opening/closing quotes that is
        immediately followed by ':', so keys like
            'Volvo China Open Day 4': [...]
        are extracted whole rather than being split on whitespace.

    Pass 2 — unquoted single-token dict keys and scalar values (legacy):
        Handles the original patterns such as
            'm30351': [...],  {key: {...}},  MOVIE_ID: 53

    Pass 3 — plain comma-separated IDs (no dict structure).
    """
    if not asset_ids_text:
        return []
    text = str(asset_ids_text).strip()
    ids = []

    def _is_valid_candidate(candidate):
        """Return True if *candidate* should be kept as an asset ID.
        Digits are NOT required — identifiers may be purely alphabetic with spaces.
        Skip-word sets filter out known field names and status tokens.
        Candidates that contain the word "Nodes" are always module/root labels,
        never real asset IDs, and are rejected here.
        """
        if not candidate:
            return False
        # "Live Feeds Nodes", "JSON Nodes", "Schedule Nodes" etc. are module labels
        if re.search(r'\bNodes?\b', candidate, re.IGNORECASE):
            return False
        norm = candidate.lower().replace(' ', '').replace('_', '')
        return norm not in _FIELD_SKIP_WORDS and norm not in _ID_SKIP_WORDS

    # Pass 1: quoted dict keys — captures full key including any spaces
    #   {'Volvo China Open Day 4': ['rating']}  →  Volvo China Open Day 4
    #   {'Magical Kenya Open': ['releaseDate']}  →  Magical Kenya Open
    #   {'m30351': ['thumbnail']}                →  m30351
    for m in re.finditer(r"""['"]([^'"]+)['"]\s*:""", text):
        candidate = m.group(1).strip()
        if _is_valid_candidate(candidate):
            ids.append(candidate)

    # Pass 2 (original): unquoted single-token keys / scalar values
    #   e.g.  {m30351: [...]}  or  key: 53
    # Digit is required here to prevent structural module-label tokens such as
    # "Nodes", "JSON", "Live", "Feeds" from being mistaken for asset IDs.
    # Those tokens are unquoted multi-word keys and always lack digits, while
    # real unquoted asset IDs (m30351, BellMedia-2931555) always contain digits.
    if not ids:
        for m in re.finditer(r"""['\"]?([A-Za-z0-9][\w\-]*)['\"]?\s*:\s*[\[{'"\d]""", text):
            candidate = m.group(1).strip()
            if (re.search(r'\d', candidate)
                    and re.match(r'^[A-Za-z0-9][\w\-]*$', candidate)
                    and _is_valid_candidate(candidate)):
                ids.append(candidate)

    # Pass 3: plain comma-separated list with no brace/bracket structure
    if not ids:
        for token in re.split(r'[,\n]+', text):
            token = token.strip().strip("'\" ")
            if token and re.match(r'^[A-Za-z0-9][\w\- ]*$', token) and _is_valid_candidate(token):
                ids.append(token)

    # Deduplicate preserving first-seen order
    seen: set = set()
    result = []
    for i in ids:
        if i not in seen:
            seen.add(i)
            result.append(i)
    return result


def _parse_asset_field_mapping(asset_ids_text):
    """Parse the Asset IDs column into an OrderedDict of {asset_id: [fields]}
    when it contains per-asset missing-field mappings such as:

        {'20QFAZ4SO2KREXLT8T8JV': ['releaseDate']},
        {'20QF8QRN2P7P37RUZ83H6': ['thumbnail', 'releaseDate']}

    Returns an OrderedDict on success, or None when the format is not
    recognised (so callers fall back to the shared-issue-text path).
    """
    text = str(asset_ids_text or '').strip()
    # Quick guard: must look like it has dict-with-list structure
    if not text or '[' not in text or ':' not in text:
        return None
    try:
        # The column holds comma-separated single-key dicts; wrap in [] to eval.
        items = ast.literal_eval(f'[{text}]')
        if not isinstance(items, list) or not items:
            return None
        result = OrderedDict()
        for item in items:
            if not isinstance(item, dict):
                return None          # mixed structure — bail out safely
            for k, v in item.items():
                if isinstance(v, list) and all(isinstance(f, str) for f in v):
                    key_str = str(k)
                    # Reject module/root-level labels such as "Live Feeds Nodes",
                    # "JSON Nodes", "Schedule Nodes" — they are never asset IDs.
                    if re.search(r'\bNodes?\b', key_str, re.IGNORECASE):
                        return None
                    key_norm = key_str.lower().replace(' ', '').replace('_', '')
                    if key_norm in _FIELD_SKIP_WORDS or key_norm in _ID_SKIP_WORDS:
                        return None
                    result[key_str] = v
                else:
                    return None      # unexpected value type — bail out safely
        return result if result else None
    except Exception:
        return None


def _normalize_issue_text(issue: str) -> str:
    """Presentation-only normalisation applied when rendering issues in the
    Failure Summary table.  Never touches validators, Excel data, or any
    other part of the report pipeline.
    """
    import re as _re_local

    # 1. "for one or more [type] assets" → "for the asset"
    issue = _re_local.sub(
        r'\bfor one or more (?:series |episode |movie |TV special |short-form video |content )?assets?\b',
        'for the asset',
        issue, flags=_re_local.IGNORECASE,
    )

    # 2. "thumbnail value is missing" → "thumbnail image is missing"
    issue = issue.replace('thumbnail value is missing', 'thumbnail image is missing')

    # 3. Leading count in description/title length messages
    #    "8 description(s) exceed …" → "description exceeds …"
    issue = _re_local.sub(
        r'\b\d+\s+description\(s\)\s+exceed\b',
        'description exceeds',
        issue, flags=_re_local.IGNORECASE,
    )
    issue = _re_local.sub(
        r'\b\d+\s+title\(s\)\s+exceed\b',
        'title exceeds',
        issue, flags=_re_local.IGNORECASE,
    )

    # 3b. Leading count in year/version label messages
    #    "82 title(s) contain a year or version label" → "Title contains a year or version label."
    issue = _re_local.sub(
        r'\b\d+\s+title\(s\)\s+contain\s+a\s+year\s+or\s+version\s+label\.?',
        'Title contain year or version label.',
        issue, flags=_re_local.IGNORECASE,
    )

    # 4. "N thumbnail(s) are not 1920x1080 resolution"
    #    → "thumbnail does not have the required 1920×1080 resolution"
    issue = _re_local.sub(
        r'\b\d+\s+thumbnail\(s\)\s+are\s+not\s+1920[xX×]1080\s+resolution\b',
        'thumbnail does not have the required 1920×1080 resolution',
        issue, flags=_re_local.IGNORECASE,
    )

    # 5. "Thumbnail not having proper resolution: (W, H)"
    #    → "thumbnail resolution is invalid (W×H)"
    issue = _re_local.sub(
        r'Thumbnail not having proper resolution:\s*\((\d+),\s*(\d+)\)',
        lambda m: f'thumbnail resolution is invalid ({m.group(1)}×{m.group(2)})',
        issue, flags=_re_local.IGNORECASE,
    )

    return issue


def _should_use_asset_field_mapping(scenario: str, issue_text: str) -> bool:
    """Return True only when the row is a genuine missing-field scenario where
    per-asset field granularity is meaningful.

    Accepted patterns:
    - Scenario matches "Check Availability of … Keys" or
      "Check Availability of … Node Value"
    - Issue text explicitly contains both "required field" and "missing"

    This prevents diagnostic payloads (schedule times, matched title strings,
    etc.) from being misinterpreted as missing-field lists.
    """
    if re.search(
        r'check\s+availability\s+of\s+.+\s+(?:keys|node\s+value)',
        scenario,
        re.IGNORECASE,
    ):
        return True
    it = issue_text.lower()
    return 'required field' in it and 'missing' in it


def _group_failure_rows(rows):
    """Group failed Excel rows into two OrderedDicts shared by the HTML table
    and the embedded Excel export — ensuring both always match.

    Returns
    -------
    asset_groups  : OrderedDict  key=asset_id    → {'modules': [...], 'issues': [...]}
    module_groups : OrderedDict  key=module_label → {'issues': [...]}
    """
    failed_rows = [r for r in rows if _norm(r.get('Status', '')) == 'Failed']
    asset_groups: OrderedDict  = OrderedDict()
    module_groups: OrderedDict = OrderedDict()

    for row in failed_rows:
        module_label  = _clean_module_label(row.get('Module', ''))
        issue_summary = str(row.get('Issue Summary', '') or '').strip()
        scenario      = str(row.get('Scenario',      '') or '').strip()
        issue_text    = issue_summary if issue_summary else scenario

        asset_ids_raw   = str(row.get('Asset IDs', '') or '').strip()
        asset_field_map = (
            _parse_asset_field_mapping(asset_ids_raw)
            if _should_use_asset_field_mapping(scenario, issue_text)
            else None
        )

        if asset_field_map:
            # Per-asset granularity path — each asset gets its own field list.
            # Preserve the module-validation prefix already present in the
            # Issue Summary (e.g. "Movies validation:") so wording is consistent.
            _pm = re.match(r'^(.+?validation:)\s*', issue_text, re.IGNORECASE)
            _prefix = (_pm.group(1) + ' ') if _pm else f'{module_label} validation: '

            for aid, fields in asset_field_map.items():
                # Deduplicate while preserving original order
                field_str = ', '.join(dict.fromkeys(str(f) for f in fields))
                per_asset_issue = _normalize_issue_text(
                    f'{_prefix}required field "{field_str}" is missing for the asset.'
                )
                if aid not in asset_groups:
                    asset_groups[aid] = {'modules': [], 'issues': []}
                if module_label and module_label not in asset_groups[aid]['modules']:
                    asset_groups[aid]['modules'].append(module_label)
                if per_asset_issue and per_asset_issue not in asset_groups[aid]['issues']:
                    asset_groups[aid]['issues'].append(per_asset_issue)

        else:
            # Existing path — shared issue text across all extracted asset IDs.
            asset_ids     = _extract_asset_ids_for_grouping(asset_ids_raw)
            display_issue = _normalize_issue_text(issue_text) if issue_text else issue_text

            if asset_ids:
                for aid in asset_ids:
                    if aid not in asset_groups:
                        asset_groups[aid] = {'modules': [], 'issues': []}
                    if module_label and module_label not in asset_groups[aid]['modules']:
                        asset_groups[aid]['modules'].append(module_label)
                    if display_issue and display_issue not in asset_groups[aid]['issues']:
                        asset_groups[aid]['issues'].append(display_issue)
            else:
                if module_label not in module_groups:
                    module_groups[module_label] = {'issues': []}
                if display_issue and display_issue not in module_groups[module_label]['issues']:
                    module_groups[module_label]['issues'].append(display_issue)

    return asset_groups, module_groups


def _render_failure_html_card(asset_groups, module_groups):
    """Render the Failure Summary HTML card from pre-grouped data.

    Accepts the same (asset_groups, module_groups) produced by
    _group_failure_rows so the HTML table and the embedded Excel always match.
    """
    blurb = (
        '<div class="tab-blurb">'
        'Below are list of Asset ID&#39;s and respective issues.'
        '</div>'
    )

    if not asset_groups and not module_groups:
        return (
            '<div class="card">'
            + blurb
            + '<div class="fs-none">No failures recorded.</div>'
            '</div>'
        )

    thead = (
        '<thead><tr>'
        '<th class="fs-col-id">Asset ID</th>'
        '<th class="fs-col-field">Field / Root Property</th>'
        '<th class="fs-col-issues">Issues</th>'
        '</tr></thead>'
    )

    def _issues_html(issues):
        if len(issues) > 1:
            return (
                '<ul style="margin:0;padding-left:16px;line-height:1.7;">'
                + ''.join(f'<li>{_esc(i)}</li>' for i in issues)
                + '</ul>'
            )
        return _esc(issues[0]) if issues else ''

    tbody_rows = ''

    # Module-level rows (no asset ID) rendered first so structural failures
    # appear at the top of the table, above per-asset failures.
    for module_label, data in module_groups.items():
        tbody_rows += (
            f'<tr>'
            f'<td class="fs-col-id"></td>'
            f'<td class="fs-col-field">{_esc(module_label)}</td>'
            f'<td class="fs-col-issues">{_issues_html(data["issues"])}</td>'
            f'</tr>'
        )

    for aid, data in asset_groups.items():
        tbody_rows += (
            f'<tr>'
            f'<td class="fs-col-id">{_esc(aid)}</td>'
            f'<td class="fs-col-field">{_esc(", ".join(data["modules"]))}</td>'
            f'<td class="fs-col-issues">{_issues_html(data["issues"])}</td>'
            f'</tr>'
        )

    return (
        '<div class="card">'
        + blurb
        + f'<table class="fs-table">{thead}<tbody>{tbody_rows}</tbody></table>'
        '</div>'
    )


def _failure_summary_excel_b64(asset_groups, module_groups, channel_name):
    """Build an in-memory Excel workbook from failure-summary groups.

    Uses the identical grouped data as the HTML table so both always match.
    Returns (base64_string, suggested_download_filename).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Failure Summary'

    ws.append(['Asset ID', 'Field / Root Property', 'Issues'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for aid, data in asset_groups.items():
        ws.append([aid, ', '.join(data['modules']), '\n'.join(data['issues'])])

    for module_label, data in module_groups.items():
        ws.append(['', module_label, '\n'.join(data['issues'])])

    for row in ws.iter_rows(min_row=2, max_col=3):
        row[2].alignment = Alignment(wrap_text=True)
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    b64_str = base64.b64encode(buf.getvalue()).decode('ascii')

    safe_ch  = re.sub(r'[^\w\-]', '_', str(channel_name or 'report')).strip('_') or 'report'
    safe_ts  = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'{safe_ch}_failure summary_{safe_ts}.xlsx'

    return b64_str, filename


def _full_report_excel_b64(excel_path, channel_name):
    """Embed the final Validation_Output Excel (all statuses) for browser download.

    Returns (base64_string, suggested_download_filename), or ('', '') if unreadable.
    """
    if not excel_path or not os.path.isfile(excel_path):
        return '', ''

    try:
        with open(excel_path, 'rb') as fh:
            b64_str = base64.b64encode(fh.read()).decode('ascii')
    except Exception as exc:
        print(f'WARNING: _full_report_excel_b64 — could not read Excel: {exc}')
        return '', ''

    safe_ch = re.sub(r'[^\w\-]', '_', str(channel_name or 'report')).strip('_') or 'report'
    safe_ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    # Prefer original basename when available for traceability.
    base = os.path.splitext(os.path.basename(excel_path))[0] or 'full_report'
    filename = f'{safe_ch}_full_report_{base}_{safe_ts}.xlsx'
    return b64_str, filename


def _group_failure_rows_from_updated_summary_list(updated_summary_list):
    """Build (asset_groups, module_groups) from the pre-filtered updated_summary_list.

    Each entry must have:
        'Asset_ID'      – asset identifier; empty/blank → module-level row
        'Module'        – field / root property label
        'Issue Summary' – human-readable issue text

    Returns the same (asset_groups, module_groups) tuple expected by
    _render_failure_html_card and _failure_summary_excel_b64.
    """
    asset_groups: OrderedDict  = OrderedDict()
    module_groups: OrderedDict = OrderedDict()

    for entry in updated_summary_list:
        asset_id     = str(entry.get('Asset_ID', entry.get('Asset ID', '')) or '').strip()
        module_label = str(entry.get('Module',        '') or '').strip()
        issue_text   = str(entry.get('Issue Summary', '') or '').strip()

        if asset_id:
            if asset_id not in asset_groups:
                asset_groups[asset_id] = {'modules': [], 'issues': []}
            if module_label and module_label not in asset_groups[asset_id]['modules']:
                asset_groups[asset_id]['modules'].append(module_label)
            if issue_text and issue_text not in asset_groups[asset_id]['issues']:
                asset_groups[asset_id]['issues'].append(issue_text)
        else:
            if module_label not in module_groups:
                module_groups[module_label] = {'issues': []}
            if issue_text and issue_text not in module_groups[module_label]['issues']:
                module_groups[module_label]['issues'].append(issue_text)

    return asset_groups, module_groups


def _build_failure_summary_card(rows):
    """Backward-compatible entry-point: groups rows then renders the HTML card."""
    ag, mg = _group_failure_rows(rows)
    return _render_failure_html_card(ag, mg)


# ---------------------------------------------------------------------------
# Tab builders — Complete / Failed / Grouped Failed
# ---------------------------------------------------------------------------

_COMPLETE_COLUMNS = (
    ('S.No', 'tc-col-sno', 56),
    ('Module', 'tc-col-module', 120),
    ('Scenario', 'tc-col-scenario', 240),
    ('Expected Results', 'tc-col-expected', 220),
    ('Status', 'tc-col-status', 110),
    ('Issue Summary', 'tc-col-issue', 240),
    ('Asset IDs', 'tc-col-assets', 200),
)

_ASSET_PREVIEW_LIMIT = 120
_GROUPED_ASSET_PREVIEW_COUNT = 5


def _filter_pill(count, label, filter_key, dot, text, bg, active=False):
    active_cls = ' active-filter' if active else ''
    return (
        f'<button type="button" class="tc-filter-pill{active_cls}" '
        f'data-filter="{_esc(filter_key)}" '
        f'style="background:{bg};color:{text};">'
        f'<span class="stat-dot" style="background:{dot};"></span>'
        f'{count} {label}</button>'
    )


def _render_asset_ids_cell(raw_assets):
    """Preview + …more expand + copy-full for Asset IDs column.

    Preview length is adjusted in the browser when the column is resized.
    """
    full = str(raw_assets or '').strip()
    if not full:
        return ''

    needs_more = len(full) > _ASSET_PREVIEW_LIMIT
    preview = full[:_ASSET_PREVIEW_LIMIT] + ('…' if needs_more else '')
    full_attr = _html.escape(full, quote=True)
    more_style = '' if needs_more else ' style="display:none"'

    return (
        f'<div class="tc-assets-cell" data-full="{full_attr}">'
        f'<div class="tc-assets-preview">{_esc(preview)}</div>'
        f'<div class="tc-assets-full">{_esc(full)}</div>'
        f'<div class="tc-assets-actions">'
        f'<button type="button" class="tc-copy-btn" data-copy="{full_attr}">Copy</button>'
        f'<button type="button" class="tc-more-btn" data-more="1"{more_style}>…more</button>'
        f'</div>'
        f'</div>'
    )


def _render_complete_test_cases_panel(rows, counts):
    """Tab 1: full Validation_Output / Excel rows with status badges."""
    blurb = (
        '<div class="tab-blurb">'
        'Below are the complete End-to-End EPG execution results are listed.'
        '</div>'
    )
    kpi = (
        '<div class="tc-kpi-row" id="tc-status-filters">'
        + _filter_pill(counts.get('Passed', 0), 'Passed', 'Passed',
                       '#22C55E', '#166534', '#DCFCE7')
        + _filter_pill(counts.get('Failed', 0), 'Failed', 'Failed',
                       '#EF4444', '#991B1B', '#FEE2E2')
        + _filter_pill(counts.get('Not Tested', 0), 'Not Tested', 'Not Tested',
                       '#3B82F6', '#1E40AF', '#DBEAFE')
        + _filter_pill(counts.get('Observation', 0), 'Observation', 'Observation',
                       '#F59E0B', '#854D0E', '#FEF9C3')
        + _filter_pill(counts.get('Not Applicable', 0), 'Not Applicable', 'Not Applicable',
                       '#94A3B8', '#475569', '#F1F5F9')
        + _filter_pill(len(rows), 'Total', 'ALL',
                       '#64748B', '#475569', '#F8FAFC', active=True)
        + '</div>'
    )

    if not rows:
        return (
            blurb
            + kpi
            + '<div class="tc-empty">No test cases recorded.</div>'
        )

    colgroup = (
        '<colgroup>'
        + ''.join(
            f'<col class="{css}" data-default-width="{width}" style="width:{width}px;">'
            for _col, css, width in _COMPLETE_COLUMNS
        )
        + '</colgroup>'
    )
    thead = (
        '<thead><tr>'
        + ''.join(
            f'<th class="{css}">'
            f'<div class="tc-th-inner">'
            f'<span class="tc-th-label">{_esc(col)}</span>'
            f'<span class="tc-col-resizer" role="separator" aria-orientation="vertical" '
            f'aria-label="Resize {_esc(col)} column" title="Drag to resize · double-click to reset"></span>'
            f'</div>'
            f'</th>'
            for col, css, _width in _COMPLETE_COLUMNS
        )
        + '</tr></thead>'
    )

    body_rows = []
    for r in rows:
        status = _norm(r.get('Status', ''))
        badge = (
            f'<span class="badge badge-{_cls(status)}">{_esc(status)}</span>'
        )
        issue = _truncate(r.get('Issue Summary', '') or '', 500)
        assets_html = _render_asset_ids_cell(r.get('Asset IDs', '') or '')
        body_rows.append(
            f'<tr data-status="{_esc(status)}">'
            f'<td class="tc-col-sno">{_esc(r.get("S.No", ""))}</td>'
            f'<td class="tc-col-module">{_esc(r.get("Module", ""))}</td>'
            f'<td class="tc-col-scenario">{_esc(r.get("Scenario", ""))}</td>'
            f'<td class="tc-col-expected">{_esc(_truncate(r.get("Expected Results", "") or "", 300))}</td>'
            f'<td class="tc-col-status">{badge}</td>'
            f'<td class="tc-col-issue">{_esc(issue)}</td>'
            f'<td class="tc-col-assets">{assets_html}</td>'
            '</tr>'
        )

    return (
        blurb
        + kpi
        + '<div class="tc-table-wrap">'
        f'<table class="tc-table" id="tc-complete-table">'
        f'{colgroup}{thead}<tbody>{"".join(body_rows)}</tbody></table>'
        '</div>'
    )


def _group_failed_rows_by_issue_summary(rows):
    """Group Excel Failed rows by Issue Summary (OrderedDict preserves first-seen order)."""
    groups = OrderedDict()
    for r in rows:
        if _norm(r.get('Status', '')) != 'Failed':
            continue
        issue = str(r.get('Issue Summary', '') or '').strip()
        if not issue:
            issue = str(r.get('Scenario', '') or '').strip() or 'Unknown Issue'
        groups.setdefault(issue, []).append(r)
    return groups


def _extract_asset_ids_for_grouped_tab(asset_ids_text, module):
    """Extract Asset ID keys from the Excel Asset IDs column.

    Schedule module shape (top-level keys are asset IDs):
        {asset_id: [date, start, …]}, {asset_id2: […]}

    Other modules shape (date → list of {asset_id: …} dicts):
        {date: [{asset_id: […]}, {asset_id2: […]}]}, …
    """
    ids = []
    seen = set()
    text = str(asset_ids_text or '').strip()
    if not text:
        return ids

    try:
        items = ast.literal_eval(f'[{text}]')
    except Exception:
        return ids

    if not isinstance(items, list):
        return ids

    is_schedule = str(module or '').strip() == 'Schedule'

    for item in items:
        if not isinstance(item, dict):
            continue

        if is_schedule:
            for key in item.keys():
                key_str = str(key)
                if key_str and key_str not in seen:
                    seen.add(key_str)
                    ids.append(key_str)
            continue

        # Non-Schedule: values are lists of dicts; take keys of those dicts.
        for _date, inner_list in item.items():
            if not isinstance(inner_list, list):
                # Fallback: if value is a dict, treat its keys as asset IDs.
                if isinstance(inner_list, dict):
                    for key in inner_list.keys():
                        key_str = str(key)
                        if key_str and key_str not in seen:
                            seen.add(key_str)
                            ids.append(key_str)
                continue
            for entry in inner_list:
                if isinstance(entry, dict):
                    for key in entry.keys():
                        key_str = str(key)
                        if key_str and key_str not in seen:
                            seen.add(key_str)
                            ids.append(key_str)

    return ids


def _render_grouped_asset_ids_body(asset_ids):
    """Render Asset ID list for grouped tab; preview first 5, …more + copy when more."""
    id_count = len(asset_ids)
    if not asset_ids:
        return '<div class="gf-empty" style="padding:10px 14px;">No Asset IDs recorded.</div>'

    full_text = '\n'.join(asset_ids)
    full_attr = _html.escape(full_text, quote=True)
    needs_more = id_count > _GROUPED_ASSET_PREVIEW_COUNT
    preview_text = '\n'.join(asset_ids[:_GROUPED_ASSET_PREVIEW_COUNT])

    actions = (
        f'<div class="tc-assets-actions">'
        f'<button type="button" class="tc-copy-btn" data-copy="{full_attr}">Copy</button>'
    )
    if needs_more:
        actions += (
            '<button type="button" class="tc-more-btn" data-more="1">…more</button>'
        )
    actions += '</div>'

    return (
        f'<div class="gf-assets-wrap">'
        f'<div class="gf-assets-header">Asset IDs ({id_count})</div>'
        f'<div class="tc-assets-cell">'
        f'<div class="tc-assets-preview gf-asset-id">{_esc(preview_text)}</div>'
        f'<div class="tc-assets-full gf-asset-id">{_esc(full_text)}</div>'
        + actions
        + '</div>'
        '</div>'
    )


def _render_grouped_failed_cases_panel(rows):
    """Tab 3: Failed cases accordion grouped by Issue Summary; body shows Asset IDs only.

    Groups are sorted descending by unique Asset ID count.
    """
    groups = _group_failed_rows_by_issue_summary(rows)
    header = ''
    blurb = (
        '<div class="tab-blurb">'
        'Below the list of Issues and respective Asset ID&#39;s'
        '</div>'
    )

    if not groups:
        return header + blurb + '<div class="gf-empty">No failures recorded.</div>'

    prepared = []
    for issue_summary, failed_rows in groups.items():
        asset_ids = []
        seen = set()
        for r in failed_rows:
            for aid in _extract_asset_ids_for_grouped_tab(
                r.get('Asset IDs', ''),
                r.get('Module', ''),
            ):
                if aid not in seen:
                    seen.add(aid)
                    asset_ids.append(aid)
        prepared.append((issue_summary, failed_rows, asset_ids))

    # Descending by unique Asset ID count; tie-break by issue text.
    prepared.sort(key=lambda item: (-len(item[2]), item[0].lower()))

    items = []
    for issue_summary, failed_rows, asset_ids in prepared:
        id_count = len(asset_ids)
        open_attr = ' open' if id_count <= 3 and id_count > 0 else ''

        if asset_ids:
            body_html = _render_grouped_asset_ids_body(asset_ids)
        else:
            body_html = '<div class="gf-empty" style="padding:10px 14px;">No Asset IDs recorded.</div>'

        items.append(
            f'<details class="gf-item"{open_attr}>'
            f'<summary class="gf-summary">'
            f'<span class="chevron">&#9658;</span>'
            f'<span class="gf-sc-name">{_esc(issue_summary)}</span>'
            f'<span class="gf-count">{id_count} asset{"s" if id_count != 1 else ""}</span>'
            f'</summary>'
            f'<div class="gf-body">{body_html}</div>'
            f'</details>'
        )

    return header + blurb + ''.join(items)


def _render_tab_shell(complete_html, failed_html, grouped_html):
    """Tab navigation + three panels. Default active tab = Failed Cases."""
    return (
        '<div class="tab-bar">'
        '<div class="tab-bar-tabs" role="tablist">'
        '<button type="button" class="tab-btn" data-tab="complete" role="tab"'
        ' aria-selected="false">Test Case + Results Section</button>'
        '<button type="button" class="tab-btn active" data-tab="failed" role="tab"'
        ' aria-selected="true">Failures Grouped under Asset IDs</button>'
        '<button type="button" class="tab-btn" data-tab="grouped" role="tab"'
        ' aria-selected="false">Asset IDs Grouped under Failures</button>'
        '</div>'
        '<div class="tab-bar-actions">'
        '<button type="button" class="dl-btn dl-btn-tab tab-bar-dl-hidden" id="dl-btn-complete"'
        ' onclick="downloadFullReport()">Download as Excel</button>'
        '<button type="button" class="dl-btn dl-btn-tab" id="dl-btn-failed"'
        ' onclick="downloadFailureSummary()">Download Failure Summary</button>'
        '</div>'
        '</div>'
        '<div class="tab-panels">'
        f'<div class="tab-panel" id="tab-complete" role="tabpanel">{complete_html}</div>'
        f'<div class="tab-panel active" id="tab-failed" role="tabpanel">{failed_html}</div>'
        f'<div class="tab-panel" id="tab-grouped" role="tabpanel">{grouped_html}</div>'
        '</div>'
    )


_TAB_SCRIPT = """
<script>
function initReportTabs(){
  var buttons = document.querySelectorAll('.tab-btn');
  var panels = document.querySelectorAll('.tab-panel');
  var dlComplete = document.getElementById('dl-btn-complete');
  var dlFailed = document.getElementById('dl-btn-failed');
  if(!buttons.length) return;

  function syncDownloadButtons(target){
    if(dlComplete){
      dlComplete.classList.toggle('tab-bar-dl-hidden', target !== 'complete');
    }
    if(dlFailed){
      dlFailed.classList.toggle('tab-bar-dl-hidden', target !== 'failed');
    }
  }

  buttons.forEach(function(btn){
    btn.addEventListener('click', function(){
      var target = btn.getAttribute('data-tab');
      buttons.forEach(function(b){
        b.classList.toggle('active', b === btn);
        b.setAttribute('aria-selected', b === btn ? 'true' : 'false');
      });
      panels.forEach(function(p){
        p.classList.toggle('active', p.id === 'tab-' + target);
      });
      syncDownloadButtons(target);
    });
  });
  // Default active tab is Failed Cases
  syncDownloadButtons('failed');
}

function initCompleteFilters(){
  var pills = document.querySelectorAll('#tc-status-filters .tc-filter-pill');
  var rows = document.querySelectorAll('#tc-complete-table tbody tr');
  if(!pills.length || !rows.length) return;

  function applyFilter(filter){
    pills.forEach(function(p){
      p.classList.toggle('active-filter', p.getAttribute('data-filter') === filter);
    });
    rows.forEach(function(row){
      var status = row.getAttribute('data-status') || '';
      var show = (filter === 'ALL' || status === filter);
      row.classList.toggle('tc-row-hidden', !show);
    });
  }

  pills.forEach(function(pill){
    pill.addEventListener('click', function(){
      applyFilter(pill.getAttribute('data-filter') || 'ALL');
    });
  });
}

function initResizableCompleteTable(){
  var table = document.getElementById('tc-complete-table');
  if(!table) return;
  var wrap = table.closest('.tc-table-wrap');
  var cols = table.querySelectorAll('colgroup col');
  var ths = table.querySelectorAll('thead th');
  if(!cols.length || !ths.length) return;

  var ASSETS_COL = 6;
  // Near-zero floor so drag-left can go almost to the edge (grip still usable)
  var MIN_COL = 16;
  var widths = [];
  cols.forEach(function(col){
    var w = parseFloat(col.style.width) || parseFloat(col.getAttribute('data-default-width') || '120');
    widths.push(Math.max(MIN_COL, w));
  });

  function decodeAttr(escaped){
    var ta = document.createElement('textarea');
    ta.innerHTML = escaped || '';
    return ta.value;
  }

  function refreshAssetPreviews(colWidth){
    var hideActions = colWidth < 72;
    var charBudget = Math.max(4, Math.floor((colWidth - (hideActions ? 8 : 28)) / 6.5));
    table.querySelectorAll('td.tc-col-assets .tc-assets-cell').forEach(function(cell){
      var actions = cell.querySelector('.tc-assets-actions');
      if(actions){
        actions.classList.toggle('tc-actions-hidden', hideActions);
      }
      if(cell.classList.contains('expanded')) return;
      var full = decodeAttr(cell.getAttribute('data-full') || '');
      var previewEl = cell.querySelector('.tc-assets-preview');
      var moreBtn = cell.querySelector('.tc-more-btn');
      if(!previewEl) return;
      if(full.length > charBudget){
        previewEl.textContent = full.slice(0, charBudget) + '…';
        if(moreBtn && !hideActions){
          moreBtn.style.display = '';
          moreBtn.textContent = '…more';
        }
      } else {
        previewEl.textContent = full;
        if(moreBtn) moreBtn.style.display = 'none';
      }
    });
  }

  function syncTableWidth(){
    var total = 0;
    for(var i = 0; i < widths.length; i++) total += widths[i];
    table.style.width = Math.round(total) + 'px';
    table.style.minWidth = Math.round(total) + 'px';
  }

  function applyCol(colIndex){
    var next = widths[colIndex];
    var px = next + 'px';
    var pad = next < 40 ? '4px 2px' : (next < 70 ? '6px 4px' : '');

    var col = cols[colIndex];
    if(col){
      col.style.width = px;
      col.style.minWidth = px;
      col.style.maxWidth = px;
    }
    var th = ths[colIndex];
    if(th){
      th.style.width = px;
      th.style.minWidth = px;
      th.style.maxWidth = px;
      if(pad) th.style.padding = pad;
      else th.style.padding = '';
    }
    var body = table.tBodies[0];
    if(body){
      for(var r = 0; r < body.rows.length; r++){
        var cell = body.rows[r].cells[colIndex];
        if(!cell) continue;
        cell.style.width = px;
        cell.style.minWidth = px;
        cell.style.maxWidth = px;
        if(pad) cell.style.padding = pad;
        else cell.style.padding = '';
      }
    }
    if(colIndex === ASSETS_COL){
      refreshAssetPreviews(next);
    }
  }

  function setColWidthExcel(colIndex, desiredPx){
    var desired = Math.round(desiredPx);
    var last = widths.length - 1;

    if(colIndex >= last){
      // Last column: grow/shrink table freely down to MIN_COL
      widths[colIndex] = Math.max(MIN_COL, desired);
      applyCol(colIndex);
      syncTableWidth();
      return;
    }

    // Middle columns: steal/give space from the NEXT column so the divider
    // tracks the mouse and left-drag can go all the way to MIN_COL.
    var old = widths[colIndex];
    var nextOld = widths[colIndex + 1];
    var pairTotal = old + nextOld;
    var newW = Math.max(MIN_COL, Math.min(desired, pairTotal - MIN_COL));
    widths[colIndex] = newW;
    widths[colIndex + 1] = pairTotal - newW;
    applyCol(colIndex);
    applyCol(colIndex + 1);
    syncTableWidth();
  }

  function setColWidthAbsolute(colIndex, desiredPx){
    widths[colIndex] = Math.max(MIN_COL, Math.round(desiredPx));
    applyCol(colIndex);
    syncTableWidth();
  }

  // Initial layout
  for(var i = 0; i < widths.length; i++){
    applyCol(i);
  }
  syncTableWidth();

  ths.forEach(function(th, colIndex){
    var resizer = th.querySelector('.tc-col-resizer');
    if(!resizer) return;

    resizer.addEventListener('dblclick', function(ev){
      ev.preventDefault();
      ev.stopPropagation();
      var def = parseInt(cols[colIndex].getAttribute('data-default-width') || '120', 10);
      // Reset this column; for non-last, take/return space from neighbor
      if(colIndex < widths.length - 1){
        var pair = widths[colIndex] + widths[colIndex + 1];
        var w = Math.max(MIN_COL, Math.min(def, pair - MIN_COL));
        widths[colIndex] = w;
        widths[colIndex + 1] = pair - w;
        applyCol(colIndex);
        applyCol(colIndex + 1);
        syncTableWidth();
      } else {
        setColWidthAbsolute(colIndex, def);
      }
    });

    resizer.addEventListener('pointerdown', function(ev){
      if(ev.pointerType === 'mouse' && ev.button !== 0) return;
      ev.preventDefault();
      ev.stopPropagation();

      var startX = ev.clientX;
      var startW = widths[colIndex];
      var dragging = true;
      resizer.classList.add('resizing');
      document.body.classList.add('tc-col-resizing');
      try { resizer.setPointerCapture(ev.pointerId); } catch(e) {}

      function edgeScroll(clientX){
        if(!wrap) return;
        var rect = wrap.getBoundingClientRect();
        var zone = 40;
        var step = 24;
        if(clientX > rect.right - zone) wrap.scrollLeft += step;
        else if(clientX < rect.left + zone) wrap.scrollLeft -= step;
      }

      function onMove(moveEv){
        if(!dragging) return;
        setColWidthExcel(colIndex, startW + (moveEv.clientX - startX));
        edgeScroll(moveEv.clientX);
      }

      function onUp(upEv){
        if(!dragging) return;
        dragging = false;
        resizer.classList.remove('resizing');
        document.body.classList.remove('tc-col-resizing');
        try { resizer.releasePointerCapture(upEv.pointerId); } catch(e) {}
        document.removeEventListener('pointermove', onMove, true);
        document.removeEventListener('pointerup', onUp, true);
        document.removeEventListener('pointercancel', onUp, true);
      }

      document.addEventListener('pointermove', onMove, true);
      document.addEventListener('pointerup', onUp, true);
      document.addEventListener('pointercancel', onUp, true);
    });
  });
}

function initAssetIdActions(){
  document.addEventListener('click', function(ev){
    var moreBtn = ev.target.closest('.tc-more-btn');
    if(moreBtn){
      var cell = moreBtn.closest('.tc-assets-cell');
      if(!cell) return;
      var expanded = cell.classList.toggle('expanded');
      moreBtn.textContent = expanded ? '…less' : '…more';
      return;
    }
    var copyBtn = ev.target.closest('.tc-copy-btn');
    if(copyBtn){
      var text = copyBtn.getAttribute('data-copy') || '';
      // data-copy is HTML-escaped; decode via textarea trick
      var ta = document.createElement('textarea');
      ta.innerHTML = text;
      var decoded = ta.value;
      if(navigator.clipboard && navigator.clipboard.writeText){
        navigator.clipboard.writeText(decoded).then(function(){
          var prev = copyBtn.textContent;
          copyBtn.textContent = 'Copied';
          setTimeout(function(){ copyBtn.textContent = prev; }, 1200);
        }).catch(function(){
          window.prompt('Copy Asset IDs:', decoded);
        });
      } else {
        window.prompt('Copy Asset IDs:', decoded);
      }
    }
  });
}

document.addEventListener('DOMContentLoaded', function(){
  initReportTabs();
  initCompleteFilters();
  initResizableCompleteTable();
  initAssetIdActions();
});
</script>
"""


# ---------------------------------------------------------------------------
# Coverage card builder
# ---------------------------------------------------------------------------


# def _build_coverage_card():
#     """
#     Read coverage stats from genai_utils module state and return an HTML card.
#     Returns an empty string if genai_utils is not importable or GENAI is off.
#     """
#     try:
#         from utils.genai_utils import get_coverage
#         cov = get_coverage()
#     except Exception:
#         return ""
#
#     ld = cov.get("lang_detect", {})
#     oc = cov.get("ocr", {})
#
#     # Suppress the card entirely if nothing was ever enqueued/tracked
#     if (ld.get("total_enqueued", 0) == 0 and oc.get("total_candidates", 0) == 0
#             and oc.get("total_enqueued", 0) == 0):
#         return ""
#
#     def _pill(label, count, pill_class="cov-reason-pill"):
#         return f'<span class="{pill_class}">{_esc(label)} &nbsp;{count}</span>'
#
#     def _stat_table(rows):
#         trs = "".join(
#             f'<tr><td>{_esc(label)}</td><td>{val}</td></tr>'
#             for label, val in rows
#         )
#         return f'<table class="cov-table">{trs}</table>'
#
#     def _reasons_block(reasons_dict, pill_class):
#         if not reasons_dict:
#             return '<div class="cov-none">None recorded</div>'
#         pills = "".join(
#             _pill(reason, count, pill_class)
#             for reason, count in sorted(reasons_dict.items(), key=lambda x: -x[1])
#         )
#         return f'<div class="cov-reasons"><div class="cov-reasons-label">Reasons</div>{pills}</div>'
#
#     # ── LANG_DETECT section ──────────────────────────────────────────────
#     ld_rows = [
#         ("Enqueued for detection",   ld.get("total_enqueued",   0)),
#         ("Succeeded",                ld.get("total_success",    0)),
#         ("Failed (API / parse)",     ld.get("total_failed_api", 0)),
#     ]
#     ld_fail_reasons = ld.get("failure_reasons", {})
#     ld_html = (
#         '<div class="cov-section">'
#         '<div class="cov-title">'
#         '<span class="cov-title-dot" style="background:#6366F1;"></span>'
#         'Language Detection (LANG_DETECT)</div>'
#         + _stat_table(ld_rows)
#         + _reasons_block(ld_fail_reasons, "cov-reason-pill")
#         + '</div>'
#     )
#
#     # ── OCR section ──────────────────────────────────────────────────────
#     ocr_rows = [
#         ("Total assets (candidates)", oc.get("total_candidates", 0)),
#         ("Eligible (all basic checks passed)", oc.get("total_eligible",   0)),
#         ("Enqueued for OCR",          oc.get("total_enqueued",   0)),
#         ("Succeeded",                 oc.get("total_success",    0)),
#         ("Failed (API / parse)",      oc.get("total_failed_api", 0)),
#         ("Skipped (pre-OCR)",         oc.get("total_skipped",    0)),
#     ]
#     ocr_skip_reasons = oc.get("skip_reasons", {})
#     ocr_fail_reasons = oc.get("failure_reasons", {})
#     ocr_html = (
#         '<div class="cov-section">'
#         '<div class="cov-title">'
#         '<span class="cov-title-dot" style="background:#F59E0B;"></span>'
#         'OCR Title Extraction (OCR_IMAGE_TEXT)</div>'
#         + _stat_table(ocr_rows)
#         + _reasons_block(ocr_skip_reasons, "cov-skip-pill")
#         + _reasons_block(ocr_fail_reasons, "cov-reason-pill")
#         + '</div>'
#     )
#
#     card = (
#         '<div class="card">'
#         '<div class="card-label">AI Processing Coverage</div>'
#         '<div class="cov-grid">'
#         + ld_html + ocr_html +
#         '</div>'
#         '</div>'
#     )
#     return card


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def summary_report_writer(
    excel_path,
    channel_name=None,
    content_partner_name=None,
    psd=None,
    json_url=None,
    updated_summary_list=None,
):
    """
    Build an HTML summary from the Excel report at *excel_path* and save it
    alongside the Excel file.

    Parameters
    ----------
    excel_path           : str  — absolute path to the generated Excel report.
    channel_name         : str  — optional channel display name shown in the
                                  Run Configuration card.
    content_partner_name : str  — optional content partner name (shown if provided).
    psd                  : str  — optional PSD reference value (shown if provided).
    json_url             : str  — runtime JSON feed URL; overrides the global
                                  JSON_URL imported from Input.py when supplied.

    Returns
    -------
    str  – path of the generated HTML file.
    """
    # Resolve the effective feed URL: runtime parameter wins; fall back to
    # the Input.py global so callers that omit json_url keep working.
    _effective_url = json_url
    #_effective_url = json_url if json_url is not None else JSON_URL
    if not excel_path or not os.path.isfile(excel_path):
        print(f'WARNING: summary_report_writer — Excel file not found: {excel_path}')
        return None

    rows = _read_excel(excel_path)
    if not rows:
        print('WARNING: summary_report_writer — Excel file is empty, skipping HTML report.')
        return None

    # Filter out excluded scenarios for HTML rendering only; Excel is unchanged.
    visible_rows = _filter_rows(rows)

    ts       = datetime.now().strftime('%B %d, %Y  ·  %H:%M:%S')
    ts_short = datetime.now().strftime('%Y-%m-%d %H:%M')
    total    = len(visible_rows)

    counts = defaultdict(int)
    for r in visible_rows:
        counts[_norm(r.get('Status', ''))] += 1
    passed_n = counts['Passed']
    failed_n = counts['Failed']
    nt_n     = counts['Not Tested']
    obs_n    = counts['Observation']
    na_n     = counts['Not Applicable']
    pass_pct = int(passed_n / total * 100) if total else 0

    # Group by module, preserving order of first appearance
    modules = OrderedDict()
    for r in visible_rows:
        m = r.get('Module', 'Unknown') or 'Unknown'
        modules.setdefault(m, []).append(r)

    # ── Derive name tokens ────────────────────────────────────────────────
    try:
        json_name = os.path.splitext(os.path.basename(_effective_url.rstrip('/')))[0]
    except Exception:
        json_name = 'feed'
    excel_basename = os.path.splitext(os.path.basename(excel_path))[0]
    html_filename  = f'Summary-report_{json_name}_{excel_basename}.html'
    out_dir        = os.path.dirname(excel_path)
    out_path       = os.path.join(out_dir, html_filename)

    # ── Overall status pill ───────────────────────────────────────────────
    overall_status = 'Passed' if failed_n == 0 else 'Failed'
    pill_color     = '#22C55E' if overall_status == 'Passed' else '#EF4444'
    pill_bg        = 'rgba(34,197,94,.12)' if overall_status == 'Passed' else 'rgba(239,68,68,.12)'
    pill_border    = 'rgba(34,197,94,.35)' if overall_status == 'Passed' else 'rgba(239,68,68,.35)'

    excel_basename_ext = os.path.basename(excel_path)   # e.g. 1748000000000.xlsx
    json_basename_ext  = os.path.basename(_effective_url.rstrip('/'))  # e.g. 6962.json

    # ── Build HTML sections ───────────────────────────────────────────────
    # Optional metadata items — only rendered when a value was supplied
    _opt_items = ''
    if channel_name:
        _opt_items += (
            f'<div class="input-item"><label>Channel Name</label>'
            f'<span class="val">{_esc(channel_name)}</span></div>'
        )
    if content_partner_name:
        _opt_items += (
            f'<div class="input-item"><label>Content Partner</label>'
            f'<span class="val">{_esc(content_partner_name)}</span></div>'
        )
    if psd:
        _opt_items += (
            f'<div class="input-item"><label>PSD</label>'
            f'<span class="val">{_esc(psd)}</span></div>'
        )

    inputs_card = (
        '<div class="card">'
        '<div class="card-label">Run Configuration</div>'
        '<div class="input-grid">'
        f'<div class="input-item"><label>JSON Feed URL</label>'
        f'<span class="val">{_esc(_effective_url)}</span></div>'
        f'<div class="input-item"><label>Input JSON Name</label>'
        f'<span class="val">{_esc(json_basename_ext)}</span></div>'
        f'<div class="input-item"><label>Feed / Channel ID</label>'
        f'<span class="val">{_esc(json_name)}</span></div>'
        f'<div class="input-item"><label>Execution Timestamp</label>'
        f'<span class="val">{_esc(ts_short)}</span></div>'
        + _opt_items
        + '</div>'
        '</div>'
    )

    module_cards = '\n'.join(_module_card(m, r) for m, r in modules.items())

    # ── Failure Summary — compute groups once, reuse for HTML + embedded Excel ──
    if updated_summary_list:
        _fs_ag, _fs_mg = _group_failure_rows_from_updated_summary_list(updated_summary_list)
    else:
        _fs_ag, _fs_mg = _group_failure_rows(visible_rows)
    _fs_b64, _fs_filename = _failure_summary_excel_b64(_fs_ag, _fs_mg, channel_name)
    _full_b64, _full_filename = _full_report_excel_b64(excel_path, channel_name)

    # Tab panels (Failed Cases = existing Failure Summary; default active)
    complete_panel = _render_complete_test_cases_panel(visible_rows, counts)
    failed_panel   = _render_failure_html_card(_fs_ag, _fs_mg)
    grouped_panel  = _render_grouped_failed_cases_panel(visible_rows)
    tabs_html      = _render_tab_shell(complete_panel, failed_panel, grouped_panel)

    # JavaScript embedded in <head>: base64 payload + Blob download + tab switch.
    # Blob download works reliably from a standalone HTML file with no server.
    _fs_script = (
        '<script>\n'
        f'var _fsDlB64="{_fs_b64}";\n'
        f'var _fsDlName="{_fs_filename}";\n'
        f'var _fullDlB64="{_full_b64}";\n'
        f'var _fullDlName="{_full_filename}";\n'
        'function downloadFailureSummary(){\n'
        '  var bin=atob(_fsDlB64),bytes=new Uint8Array(bin.length);\n'
        '  for(var i=0;i<bin.length;i++)bytes[i]=bin.charCodeAt(i);\n'
        '  var blob=new Blob([bytes],{type:"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"});\n'
        '  var a=document.createElement("a");a.href=URL.createObjectURL(blob);a.download=_fsDlName;\n'
        '  document.body.appendChild(a);a.click();\n'
        '  setTimeout(function(){document.body.removeChild(a);URL.revokeObjectURL(a.href);},0);\n'
        '}\n'
        'function downloadFullReport(){\n'
        '  if(!_fullDlB64){alert("Full Excel report is not available.");return;}\n'
        '  var bin=atob(_fullDlB64),bytes=new Uint8Array(bin.length);\n'
        '  for(var i=0;i<bin.length;i++)bytes[i]=bin.charCodeAt(i);\n'
        '  var blob=new Blob([bytes],{type:"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"});\n'
        '  var a=document.createElement("a");a.href=URL.createObjectURL(blob);a.download=_fullDlName;\n'
        '  document.body.appendChild(a);a.click();\n'
        '  setTimeout(function(){document.body.removeChild(a);URL.revokeObjectURL(a.href);},0);\n'
        '}\n'
        '</script>\n'
        + _TAB_SCRIPT
    )

    # ── Assemble full document ────────────────────────────────────────────
    html_doc = (
        '<!DOCTYPE html>\n'
        '<html lang="en">\n'
        '<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        f'<title>Test Summary Report &mdash; {_esc(json_name)}</title>\n'
        f'<style>{_CSS}</style>\n'
        + _fs_script +
        '</head>\n'
        '<body>\n'

        # Hero
        '<div class="hero">'
        '<div class="hero-inner">'
        '<div class="hero-eyebrow">Samsung EPG Quality Report</div>'
        '<div class="hero-title">Test Summary Report</div>'
        f'<div class="hero-meta">'
        f'<span>&#128225;&nbsp;{_esc(_effective_url)}</span>'
        f'<span>&#128336;&nbsp;{_esc(ts)}</span>'
        f'</div>'
        '</div>'
        '</div>'

        # Body
        '<div class="container">'
        + inputs_card
        + tabs_html
        + '</div>'

        # Footer
        '<div class="footer">'
        f'<strong>Samsung EPG Validation Framework</strong>'
        f'&ensp;&middot;&ensp;Source: {_esc(os.path.basename(excel_path))}'
        f'&ensp;&middot;&ensp;{_esc(ts)}'
        '</div>'

        '</body>\n</html>\n'
    )

    with open(out_path, 'w', encoding='utf-8') as fh:
        fh.write(html_doc)

    print(f'INFO: HTML summary report saved → {out_path}')
    return out_path