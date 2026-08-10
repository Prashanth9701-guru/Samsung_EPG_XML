import time

import openpyxl
import logging

from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

def xlsx_report(Validation_Output, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [key for key in Validation_Output[0].keys()]
    ws.append(headers)
    logger.info(f'Xlsx Headers {headers}')
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    for cell in ws[1]:
        cell.fill = orange_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in Validation_Output:
        ws.append([str(row.get(key, '')) for key in headers])

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    light_sky_blue = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    light_yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[4]
        status_cell_value = status_cell.value.strip().lower()

        if status_cell_value in ('pass', 'passed'):
            status_cell.fill=green_fill
            status_cell.font = Font(bold=True)

        if status_cell_value in ('fail', 'failed'):
            status_cell.fill = red_fill
            status_cell.font = Font(bold=True)

        if status_cell_value in ('Not Tested', 'not tested'):
            status_cell.fill = light_sky_blue
            status_cell.font = Font(bold=True)


    MAX_WIDTH = 50  # You can adjust this (40–60 is ideal)

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, MAX_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted_width

    xlsx_path = f"{file_path}/{int(time.time()*1000)}.xlsx"
    wb.save(xlsx_path)
    return xlsx_path

